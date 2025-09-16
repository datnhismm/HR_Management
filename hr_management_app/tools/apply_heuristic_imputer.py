"""Apply heuristic imputer to an XLSX file and produce an imputed file + report.
Usage: use the project venv python to run this script.
"""

import os
import sys
from pprint import pprint

ROOT = os.path.dirname(os.path.dirname(__file__))
SRC = os.path.join(ROOT, "src")
if SRC not in sys.path:
    sys.path.insert(0, SRC)

try:
    from hr_management_app.src.ml.imputer_heuristic import (
        fit_from_records,
        predict_batch,
    )
except Exception:
    from ml.imputer_heuristic import fit_from_records, predict_batch  # type: ignore
try:
    from openpyxl import Workbook, load_workbook
except Exception as exc:
    raise SystemExit(
        "openpyxl not installed. Run: python -m pip install openpyxl"
    ) from exc

IN_FILE = os.path.join(
    os.path.dirname(os.path.dirname(__file__)), "data", "dummy_import_20k.xlsx"
)
OUT_FILE = os.path.join(
    os.path.dirname(os.path.dirname(__file__)), "data", "dummy_import_20k_imputed.xlsx"
)


def read_xlsx(path):
    wb = load_workbook(path, read_only=True)
    ws = wb.active
    if ws is None:
        raise RuntimeError("Workbook has no active sheet")
    # values_only=True returns tuples of values; the first tuple is the header row
    it = ws.iter_rows(values_only=True)
    headers_row = list(next(it))
    headers = [str(h) if h is not None else "" for h in headers_row]
    records = []
    for row in it:
        r = dict(zip(headers, row))
        records.append(r)
    return records


def write_xlsx(records, path):
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    if not records:
        wb.save(path)
        return
    headers = [str(h) for h in list(records[0].keys())]
    assert ws is not None
    ws.append(headers)
    for r in records:
        assert ws is not None
        ws.append([r.get(h) for h in headers])
    wb.save(path)


if __name__ == "__main__":
    print("Reading input...")
    recs = read_xlsx(IN_FILE)
    print(f"Read {len(recs)} rows")
    model = fit_from_records(recs)
    imputed = predict_batch(recs, model)

    # compute stats
    counts = {
        "imputed_email": 0,
        "imputed_name": 0,
        "imputed_job": 0,
        "imputed_year": 0,
    }
    for r in imputed:
        if r.get("_imputed_email"):
            counts["imputed_email"] += 1
        if r.get("_imputed_name"):
            counts["imputed_name"] += 1
        if r.get("_imputed_job"):
            counts["imputed_job"] += 1
        if r.get("_imputed_year"):
            counts["imputed_year"] += 1

    print("Imputation counts:")
    pprint(counts)
    print("Writing imputed workbook...")
    write_xlsx(imputed, OUT_FILE)
    print("Wrote", OUT_FILE)
