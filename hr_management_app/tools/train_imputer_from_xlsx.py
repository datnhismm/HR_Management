"""Train imputer models from an XLSX file and save sklearn joblib artifact.

Usage: run with the project's python executable, e.g.:
    .venv/Scripts/python.exe hr_management_app/tools/train_imputer_from_xlsx.py
"""

import os
from pprint import pprint

try:
    from hr_management_app.src.ml.imputer_ml import fit_imputer_from_records
except Exception:
    # allow running from repo root when package not installed
    from ml.imputer_ml import fit_imputer_from_records  # type: ignore
try:
    from openpyxl import load_workbook
except Exception as exc:
    raise SystemExit(
        "openpyxl not installed. Run: python -m pip install openpyxl"
    ) from exc

IN_FILE = os.path.join(
    os.path.dirname(os.path.dirname(__file__)), "data", "dummy_import_20k.xlsx"
)
OUT_PATH = os.path.join(
    os.path.dirname(os.path.dirname(__file__)), "src", "models", "imputer_model.joblib"
)


def read_xlsx(path):
    wb = load_workbook(path, read_only=True)
    ws = wb.active
    if ws is None:
        raise RuntimeError("Workbook has no active sheet")
    it = ws.iter_rows(values_only=True)
    headers = list(next(it))
    records = []
    for row in it:
        r = dict(zip(headers, row))
        # normalize keys to expected names (lowercase); coerce keys to str
        rec = {str(k).lower(): v for k, v in r.items()}
        records.append(rec)
    return records


if __name__ == "__main__":
    print("Reading", IN_FILE)
    recs = read_xlsx(IN_FILE)
    print("Loaded", len(recs), "rows")

    # filter rows with at least some fields
    print("Training sklearn models (if sklearn available)...")
    model = fit_imputer_from_records(recs, save_to=OUT_PATH)
    print("Saved model to", OUT_PATH)
    try:
        print("Model type:", model.get("type"))
        if model.get("type") == "sklearn":
            le_job = model.get("le_job")
            print("Job classes:", len(getattr(le_job, "classes_", [])))
    except Exception:
        pass

    # quick smoke predictions on a few rows with missing fields
    samples = [r for r in recs if not r.get("job_title") or not r.get("year_start")][
        :10
    ]
    if samples:
        print(
            "Running sample predictions on first",
            len(samples),
            "rows with missing values",
        )
        try:
            from hr_management_app.src.ml.imputer_ml import load_model as _load
            from hr_management_app.src.ml.imputer_ml import predict_batch
        except Exception:
            from ml.imputer_ml import load_model as _load  # type: ignore
            from ml.imputer_ml import predict_batch
        m2 = _load(OUT_PATH)
        if m2 is None:
            raise RuntimeError(f"Failed to load model from {OUT_PATH}")
        out = predict_batch(samples, m2)
        pprint(out[:5])
    else:
        print("No sample rows with missing values found for smoke test")
