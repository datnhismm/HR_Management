"""Headless preview of imputation proposals for debugging the UI preview flow.
Prints counts of proposed fields and sample proposals.
"""

import os
import sys

ROOT = os.path.dirname(os.path.dirname(__file__))
SRC = os.path.join(ROOT, "src")
if SRC not in sys.path:
    sys.path.insert(0, SRC)

try:
    from hr_management_app.src.ml.imputer_heuristic import fit_from_records
    from hr_management_app.src.ml.imputer_heuristic import predict_batch as heur_predict
    from hr_management_app.src.ml.imputer_ml import load_model, predict_batch
except Exception:
    from ml.imputer_heuristic import fit_from_records
    from ml.imputer_heuristic import predict_batch as heur_predict  # type: ignore
    from ml.imputer_ml import load_model, predict_batch  # type: ignore

IN_FILE = os.path.join(
    os.path.dirname(os.path.dirname(__file__)), "data", "dummy_import_20k.xlsx"
)


def read_xlsx_simple(path):
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True)
    ws = wb.active
    if ws is None:
        raise RuntimeError("Workbook has no active sheet")
    it = ws.iter_rows(values_only=True)
    headers_row = list(next(it))
    headers = [str(h) if h is not None else "" for h in headers_row]
    records = []
    for row in it:
        r = dict(zip(headers, row))
        # coerce keys to str and lowercase to be safe with various cell types
        rec = {str(k).lower(): v for k, v in r.items()}
        records.append(rec)
    return records


if __name__ == "__main__":
    recs = read_xlsx_simple(IN_FILE)
    print("Loaded", len(recs), "rows")

    # Clean a few rows
    sample = recs[:100]
    cleaned = [dict(r) for r in sample]

    # ML proposals
    m = load_model()
    ml_props = predict_batch(cleaned, m) if m else [{} for _ in cleaned]

    # heuristic proposals
    heur = fit_from_records(cleaned)
    h_props = heur_predict(cleaned, heur)

    # compute counts
    counts = {"ml_job": 0, "ml_year": 0, "heur_job": 0, "heur_year": 0}
    for a, b in zip(ml_props, h_props):
        if a.get("_imputed_job_conf"):
            counts["ml_job"] += 1
        if a.get("_imputed_year_pred"):
            counts["ml_year"] += 1
        if b.get("_imputed_job"):
            counts["heur_job"] += 1
        if b.get("_imputed_year"):
            counts["heur_year"] += 1

    print("Sample proposal counts (first 100 rows):")
    print(counts)
    print("Sample ML proposal example:")
    print(ml_props[:5])
    print("Sample Heuristic proposal example:")
    print(h_props[:5])
